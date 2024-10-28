

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 读取xlsx文件
file_path = 'exploration_zipcode_sort.xlsx'
df = pd.read_excel(file_path,header=None)

# 分割第二列并创建新行
split_data = df.iloc[:, 1].str.split('!!', expand=True)

# 获取第一列并插入到第五行
first_column = df.iloc[:, 0]
result = pd.concat([split_data, first_column], axis=1)

# 转置数据
result_transposed = result.transpose()

# 保存转置后的结果到新的Excel文件
output_path = 'output_file.xlsx'
result_transposed.to_excel(output_path, index=False)

# 使用openpyxl加载工作簿
wb = load_workbook(output_path)
ws = wb.active

# 合并相邻相同的单元格
for row in ws.iter_rows():
    start_col = None
    prev_value = None
    for col in range(1, len(row) + 1):
        current_value = row[col - 1].value
        if current_value == prev_value:
            # 如果当前值与之前的相同，则继续合并
            continue
        else:
            # 如果遇到不同的值，检查是否需要合并之前的单元格
            if start_col and col - start_col > 1:
                ws.merge_cells(start_row=row[0].row, start_column=start_col, 
                               end_row=row[0].row, end_column=col - 1)
            start_col = col
            prev_value = current_value

    # 检查并合并最后一组相同的单元格
    if start_col and len(row) - start_col >= 1:
        ws.merge_cells(start_row=row[0].row, start_column=start_col, 
                       end_row=row[0].row, end_column=len(row))

# 保存最终结果
wb.save(output_path)

print("文件已成功转换、处理并保存。")